"""Run read-only final delivery checks and write a concise QA report."""

from __future__ import annotations

import csv
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import nbformat
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "reports" / "delivery_qa_report.md"


@dataclass
class Check:
    name: str
    passed: bool
    evidence: str
    severity: str = "error"


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def check_required_files() -> Check:
    required = [
        "docs/project_charter.md",
        "docs/data_dictionary.md",
        "docs/metric_dictionary.md",
        "docs/optimization_log.md",
        "docs/environment_and_reproduction.md",
        "docs/interview_story.md",
        "sql/formal/00_source_profile.sql",
        "sql/formal/01_cohort_quality.sql",
        "sql/formal/02_kpi_summary_48h.sql",
        "sql/formal/03_brand_metrics_48h.sql",
        "sql/formal/04_funnel_metrics.sql",
        "sql/formal/05_window_sensitivity.sql",
        "sql/formal/06_cohort_detail_48h.sql",
        "notebooks/11_正式分析与验证.ipynb",
        "outputs/formal_delivery/eCommerce_brand_priority_formal.xlsx",
        "reports/management_report_artifact.json",
        "reports/power_bi_formal_spec.md",
        "reports/power_bi_measures.dax",
    ]
    missing = [item for item in required if not (ROOT / item).exists()]
    return Check("正式交付文件存在", not missing, "缺失：" + ", ".join(missing) if missing else f"{len(required)} 个关键文件均存在")


def check_kpi_arithmetic() -> Check:
    row = csv_rows(ROOT / "reports/data_exports/kpi_summary_48h.csv")[0]
    cohort = int(row["cohort_count"])
    counts = sum(int(row[field]) for field in ("purchase_count", "unresolved_count", "remove_count"))
    expected = (772119, 104696, 511286, 156137)
    actual = (
        cohort,
        int(row["purchase_count"]),
        int(row["unresolved_count"]),
        int(row["remove_count"]),
    )
    passed = cohort == counts and actual == expected
    return Check("A/B/C 总量与正式基准", passed, f"队列/A/B/C={actual}；A+B+C={counts}")


def check_reconciliation() -> Check:
    rows = csv_rows(ROOT / "reports/data_exports/reconciliation_summary.csv")
    errors = [row["check_name"] for row in rows if row["severity"] == "error" and row["passed"].lower() != "true"]
    warnings = [row["check_name"] for row in rows if row["severity"] == "warning" and row["passed"].lower() != "true"]
    evidence = f"检查 {len(rows)} 项；错误 {len(errors)}；已知警告 {len(warnings)}"
    if errors:
        evidence += "；错误项：" + ", ".join(errors)
    return Check("SQL/Python 跨工具对账", not errors, evidence)


def check_notebook() -> Check:
    path = ROOT / "notebooks/11_正式分析与验证.ipynb"
    notebook = nbformat.read(path, as_version=4)
    code_cells = [cell for cell in notebook.cells if cell.cell_type == "code"]
    executed = [cell for cell in code_cells if cell.execution_count is not None]
    errors = [
        output
        for cell in code_cells
        for output in cell.get("outputs", [])
        if output.get("output_type") == "error"
    ]
    passed = len(executed) == len(code_cells) and not errors
    return Check("正式 Notebook 执行状态", passed, f"代码单元 {len(code_cells)}；已执行 {len(executed)}；错误输出 {len(errors)}")


def check_excel() -> Check:
    path = ROOT / "outputs/formal_delivery/eCommerce_brand_priority_formal.xlsx"
    workbook = load_workbook(path, data_only=False, read_only=True)
    expected_sheets = {
        "01_管理层总览",
        "02_品牌优先级",
        "03_窗口敏感性",
        "04_对账与质量",
        "00_口径与参数",
        "05_KPI数据",
    }
    formula_errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "e" or (isinstance(cell.value, str) and cell.value.startswith("#")):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    missing = sorted(expected_sheets.difference(workbook.sheetnames))
    passed = not missing and not formula_errors
    evidence = f"工作表 {len(workbook.sheetnames)}；缺失 {len(missing)}；公式错误单元格 {len(formula_errors)}"
    return Check("正式 Excel 结构与公式错误", passed, evidence)


def check_excel_brand_rules() -> Check:
    """Reconcile Excel brand thresholds and classifications to the SQL export."""
    csv_data = csv_rows(ROOT / "reports/data_exports/brand_metrics_48h.csv")
    expected_threshold = float(csv_data[0]["volume_median_count"])
    expected_by_brand = {
        row["brand"]: (row["volume_band"], row["risk_band"], row["priority_quadrant"])
        for row in csv_data
    }

    path = ROOT / "outputs/formal_delivery/eCommerce_brand_priority_formal.xlsx"
    workbook = load_workbook(path, data_only=True, read_only=True)
    parameter_sheet = workbook["00_口径与参数"]
    brand_sheet = workbook["02_品牌优先级"]
    actual_threshold = parameter_sheet["B5"].value

    mismatches: list[str] = []
    seen_brands: set[str] = set()
    for row in brand_sheet.iter_rows(min_row=3, values_only=True):
        brand = row[0]
        if not brand:
            continue
        seen_brands.add(str(brand))
        actual = (row[13], row[14], row[15])
        expected = expected_by_brand.get(str(brand))
        if expected != actual:
            mismatches.append(f"{brand}:{actual}!={expected}")

    missing_brands = sorted(set(expected_by_brand).difference(seen_brands))
    threshold_matches = (
        isinstance(actual_threshold, (int, float))
        and abs(float(actual_threshold) - expected_threshold) <= 1e-9
    )
    passed = threshold_matches and not mismatches and not missing_brands
    evidence = (
        f"覆盖量阈值 Excel/SQL={actual_threshold}/{expected_threshold:g}；"
        f"品牌分类不一致 {len(mismatches)}；缺失品牌 {len(missing_brands)}"
    )
    if mismatches:
        evidence += "；示例：" + ", ".join(mismatches[:3])
    return Check("Excel 品牌阈值与分类对账", passed, evidence)


def check_sql_read_only() -> Check:
    pattern = re.compile(r"^\s*(create|insert|update|delete|drop|alter|truncate|merge)\b", re.IGNORECASE | re.MULTILINE)
    hits: list[str] = []
    for path in sorted((ROOT / "sql/formal").glob("*.sql")):
        text = path.read_text(encoding="utf-8")
        if pattern.search(text):
            hits.append(path.name)
    return Check("正式 SQL 保持只读", not hits, "未发现写操作" if not hits else "发现写操作：" + ", ".join(hits))


def check_secret_patterns() -> Check:
    files = [
        ROOT / ".env.example",
        ROOT / "README.md",
        *sorted((ROOT / "scripts").glob("*.py")),
        *sorted((ROOT / "sql/formal").glob("*.sql")),
        *sorted((ROOT / "docs").glob("*.md")),
    ]
    assignment = re.compile(r"(?i)(password|passwd|api[_-]?key|secret)\s*=\s*(['\"])([^'\"]+)\2")
    credential_url = re.compile(r"(?i)(postgres(?:ql)?://[^\s:/]+):([^@\s]+)@")
    hits: list[str] = []
    for path in files:
        text = path.read_text(encoding="utf-8")
        for match in assignment.finditer(text):
            value = match.group(3).strip()
            if value and value.lower() not in {"none", "null", "changeme", "your_password"}:
                hits.append(f"{path.relative_to(ROOT)}:{match.group(1)}")
        if credential_url.search(text):
            hits.append(f"{path.relative_to(ROOT)}:credential_url")
    return Check("正式文件未发现硬编码凭据", not hits, "未发现疑似凭据" if not hits else "疑似位置：" + ", ".join(hits))


def check_utf8() -> Check:
    files = [
        ROOT / "README.md",
        ROOT / "worklog.md",
        *sorted((ROOT / "docs").glob("*.md")),
        *sorted((ROOT / "reports").glob("*.md")),
        *sorted((ROOT / "scripts").glob("*.py")),
        *sorted((ROOT / "sql/formal").glob("*")),
    ]
    hits: list[str] = []
    for path in files:
        if not path.is_file():
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            hits.append(f"{path.relative_to(ROOT)}:非UTF-8")
            continue
        if "\ufffd" in text or "\u951f\u65a4\u62f7" in text:
            hits.append(f"{path.relative_to(ROOT)}:替换/乱码字符")
    return Check("中文与 UTF-8 抽查", not hits, f"检查 {len(files)} 个文本文件；异常 {len(hits)}" + ("；" + ", ".join(hits) if hits else ""))


def check_readme_links() -> Check:
    text = (ROOT / "README.md").read_text(encoding="utf-8")
    targets = re.findall(r"!?(?:\[[^\]]*\])\(([^)]+)\)", text)
    missing: list[str] = []
    checked = 0
    for raw_target in targets:
        target = raw_target.strip().strip("<>").split("#", 1)[0]
        if not target or re.match(r"^[a-z]+://", target, re.IGNORECASE):
            continue
        checked += 1
        if not (ROOT / target).exists():
            missing.append(target)
    return Check("README 本地链接", not missing, f"检查 {checked} 个本地链接；缺失 {len(missing)}" + ("；" + ", ".join(missing) if missing else ""))


def check_report_artifact() -> Check:
    artifact = json.loads((ROOT / "reports/management_report_artifact.json").read_text(encoding="utf-8"))
    datasets = artifact["snapshot"]["datasets"]
    sources = artifact["manifest"]["sources"]
    tables = {table for source in sources for table in source.get("query", {}).get("tables_used", [])}
    passed = (
        artifact.get("surface") == "report"
        and artifact["snapshot"].get("status") == "ready"
        and len(datasets) == 6
        and tables == {"makeup_consumer_events.dec"}
        and len(datasets["brand_metrics"]) == 97
        and len(datasets["priority_brands"]) == 20
    )
    evidence = f"状态={artifact['snapshot'].get('status')}；数据集={len(datasets)}；品牌={len(datasets['brand_metrics'])}；优先品牌={len(datasets['priority_brands'])}；源表={sorted(tables)}"
    return Check("管理层报告快照", passed, evidence)


def write_report(checks: list[Check]) -> None:
    errors = [check for check in checks if not check.passed and check.severity == "error"]
    lines = [
        "# 正式交付 QA 报告",
        "",
        f"- 总检查项：{len(checks)}",
        f"- 通过：{sum(check.passed for check in checks)}",
        f"- 错误：{len(errors)}",
        "",
        "| 检查 | 结果 | 证据 |",
        "|---|---|---|",
    ]
    for check in checks:
        status = "通过" if check.passed else "失败"
        evidence = check.evidence.replace("|", "\\|").replace("\n", " ")
        lines.append(f"| {check.name} | {status} | {evidence} |")
    lines.extend(
        [
            "",
            "## 说明",
            "",
            "本报告检查交付文件、正式口径总量、SQL/Python 对账、Notebook 执行状态、Excel 结构、Excel 品牌规则对账、SQL 只读性、凭据模式、UTF-8、本地链接和管理层报告快照。它不替代 Power BI Desktop 中的切片器交互 QA。",
            "",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    checks = [
        check_required_files(),
        check_kpi_arithmetic(),
        check_reconciliation(),
        check_notebook(),
        check_excel(),
        check_excel_brand_rules(),
        check_sql_read_only(),
        check_secret_patterns(),
        check_utf8(),
        check_readme_links(),
        check_report_artifact(),
    ]
    write_report(checks)
    for check in checks:
        print(f"[{'PASS' if check.passed else 'FAIL'}] {check.name}: {check.evidence}")
    failures = [check for check in checks if not check.passed and check.severity == "error"]
    print(f"Summary: {len(checks) - len(failures)}/{len(checks)} passed; {len(failures)} errors")
    print(f"Report: {REPORT_PATH.relative_to(ROOT)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
